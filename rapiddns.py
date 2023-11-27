# -*- coding: utf-8 -*-
import argparse
import datetime
from urllib.parse import urljoin
import httpx
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from loguru import logger

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
}


def banner():
    print('''
    ▄▄▄   ▄▄▄·  ▄▄▄·▪  ·▄▄▄▄  ·▄▄▄▄   ▐ ▄ .▄▄ · ▪        
    ▀▄ █·▐█ ▀█ ▐█ ▄███ ██▪ ██ ██▪ ██ •█▌▐█▐█ ▀. ██ ▪     
    ▐▀▀▄ ▄█▀▀█  ██▀·▐█·▐█· ▐█▌▐█· ▐█▌▐█▐▐▌▄▀▀▀█▄▐█· ▄█▀▄ 
    ▐█•█▌▐█ ▪▐▌▐█▪·•▐█▌██. ██ ██. ██ ██▐█▌▐█▄▪▐█▐█▌▐█▌.▐▌
    .▀  ▀ ▀  ▀ .▀   ▀▀▀▀▀▀▀▀• ▀▀▀▀▀• ▀▀ █▪ ▀▀▀▀▀▀▀▀ ▀█▄▀▪
                     
     @Auth: C1ph3rX13
     @Blog: https://c1ph3rx13.github.io
     @Note: 代码仅供学习使用，请勿用于其他用途  
    ''')


@logger.catch
def send_request(target: str) -> list:
    params = {
        'full': '1',
        'down': '1'
    }
    with httpx.Client(headers=headers, verify=False, timeout=10) as client:
        response = client.get(url=target, params=params)
        if response.status_code == httpx.codes.OK:
            soup = BeautifulSoup(response.text, "lxml")
            rows = soup.find("table", id="table").findAll("tr")
            items = []
            for row in rows:
                cells = row.findAll("td")
                items.append([value.text.strip() for value in cells])
            return items[1:]


@logger.catch
def subdomain(target: str) -> list:
    url = urljoin('https://rapiddns.io/subdomain/', target)
    return send_request(url)


@logger.catch
def sameip(target: str) -> list:
    url = urljoin('https://rapiddns.io/sameip/', target)
    return send_request(url)


@logger.catch
def generate_filename(action: str) -> str:
    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')
    filename_mapping = {
        'subdomain': 'RapiddnsSubdomain',
        'sameip': 'RapiddnsSameIp',
        'all': 'RapiddnsAllData'
    }
    if action in filename_mapping:
        return f'{current_time} - {filename_mapping[action]}.xlsx'
    raise ValueError('Invalid action')


@logger.catch
def write_excel(data: list, sheet_name: str, workbook: Workbook, columns: list, title_font: Font):
    if data:
        sheet = workbook.create_sheet(title=sheet_name)
        for col_num, column_title in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_num, value=column_title)
            cell.font = title_font
        for row in data:
            sheet.append(row)


if __name__ == '__main__':
    banner()

    parser = argparse.ArgumentParser(description='Rapiddns.io Data Retrieval By C1ph3rX13')
    parser.add_argument('action', choices=['subdomain', 'sameip', 'all'], help='Action to perform')
    parser.add_argument('-t', '--target', required=True, help='Target domain')
    args = parser.parse_args()

    subdomain_data = subdomain(args.target) if args.action in ['subdomain', 'all'] else []
    sameip_data = sameip(args.target) if args.action in ['sameip', 'all'] else []

    # 设置首行
    cols = ['Domain', 'Address', 'Type', 'Date']
    # 设置首行字体，颜色，字体大小
    tf = Font(bold=True, color="EE3F4D", size=20)

    # 根据参数生成文件名
    filename = generate_filename(args.action)
    wk = Workbook()

    if args.action in ['subdomain', 'all']:
        write_excel(subdomain_data, 'subdomain', wk, cols, tf)

    if args.action in ['sameip', 'all']:
        write_excel(sameip_data, 'sameip', wk, cols, tf)

    wk.remove(wk['Sheet'])  # 删除默认生成的空白工作表
    wk.save(filename)
    logger.info(f'Successfully wrote data to file: {filename}')
